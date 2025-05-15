import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

nombre_archivo = 'datos.xlsx'
#Comprobamos si el archivo ya existe
if os.path.exists(nombre_archivo):
    #Si existe, lo abrimos
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
#Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.append(['Fecha de registro', 'Fecha de incidencia','Nombre', 'Producto', 'Supervisor', 'cliente', 'Gravedad'])
#Guardar el libro de Excel
    wb.save('datos.xlsx')

#Funcion para guardar los datos en el libro de Excel
def guardar_datos():
    fecha_registro = entry_fecha_registro.get()
    fecha_incidencia = entry_fecha_incidencia.get()
    nombre = entry_nombre.get()
    producto = entry_producto.get()
    supervisor = entry_supervisor.get()
    cliente = entry_cliente.get()
    gravedad = entry_gravedad.get()

    if not fecha_registro or not fecha_incidencia or not nombre or not producto or not supervisor or not cliente or not gravedad:
        messagebox.showerror("Error", "Por favor, complete todos los campos.")
        return
    try:
        gravedad = int(gravedad)
    except ValueError:
        messagebox.showerror("Error", "La gravedad debe ser un número.")
        return
    
    #Validar el formato de la fecha
    if not re.match(r'\d{2}/\d{2}/\d{4}', fecha_registro) or not re.match(r'\d{2}/\d{2}/\d{4}', fecha_incidencia):
        messagebox.showerror("Error", "El formato de la fecha debe ser DD/MM/YYYY.")
        return
    
    ws.append([fecha_registro, fecha_incidencia, nombre, producto, supervisor, cliente, gravedad])
    wb.save(nombre_archivo)
    messagebox.showinfo("Exito", "Datos guardados correctamente.")

    #Limpiar los campos
    entry_fecha_registro.delete(0, tk.END)
    entry_fecha_incidencia.delete(0, tk.END)
    entry_nombre.delete(0, tk.END)
    entry_producto.delete(0, tk.END)
    entry_supervisor.delete(0, tk.END)
    entry_cliente.delete(0, tk.END)
    entry_gravedad.delete(0, tk.END)
    

#Crear la ventana principal
root = tk.Tk()
root.title("Formulario de Registro")
root.config(bg="#191970")
#Creando los estilos
label_style = {"bg": "#778899", "fg": "white"}
entry_style = {"bg": "#C0C0C0", "fg": "black"}

labeel_fecha_registro = tk.Label(root, text="Fecha de registro:", **label_style, anchor="w", width=20)
labeel_fecha_registro.grid(row=0, column=0, padx=10, pady=5)
entry_fecha_registro = tk.Entry(root, **entry_style)
entry_fecha_registro.grid(row=0, column=1, padx=10, pady=10)

label_fecha_incidencia = tk.Label(root, text="Fecha de incidencia:", **label_style, anchor="w", width=20)
label_fecha_incidencia.grid(row=1, column=0, padx=10, pady=5)
entry_fecha_incidencia = tk.Entry(root, **entry_style)
entry_fecha_incidencia.grid(row=1, column=1, padx=10, pady=5)

label_nombre = tk.Label(root, text="Nombre:", **label_style, anchor="w", width=20)
label_nombre.grid(row=2, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=2, column=1, padx=10, pady=5)

label_producto = tk.Label(root, text="Producto:", **label_style, anchor="w", width=20)
label_producto.grid(row=3, column=0, padx=10, pady=5)
entry_producto = tk.Entry(root, **entry_style)
entry_producto.grid(row=3, column=1, padx=10, pady=5)

label_supervisor = tk.Label(root, text="Supervisor:", **label_style, anchor="w", width=20)
label_supervisor.grid(row=4, column=0, padx=10, pady=5)
entry_supervisor = tk.Entry(root, **entry_style)
entry_supervisor.grid(row=4, column=1, padx=10, pady=5)

label_cliente = tk.Label(root, text="Cliente:", **label_style, anchor="w", width=20)
label_cliente.grid(row=5, column=0, padx=10, pady=5)
entry_cliente = tk.Entry(root, **entry_style)
entry_cliente.grid(row=5, column=1, padx=10, pady=5)

label_gravedad = tk.Label(root, text="Gravedad:", **label_style, anchor="w", width=20)
label_gravedad.grid(row=6, column=0, padx=10, pady=5)
entry_gravedad = tk.Entry(root, **entry_style)
entry_gravedad.grid(row=6, column=1, padx=10, pady=5)

#Creando el boton para guardar los datos
boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg="#6D8299", fg="white")
boton_guardar.grid(row=7, column=0, columnspan=2, pady=10, padx=10)
root.mainloop()
